import pymysql
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU
import os


def connect_to_database(host, port, user, password, database):
    """ 连接到MySQL数据库 """
    connection = pymysql.connect(host=host,
                                 port=port,
                                 user=user,
                                 password=password,
                                 database=database,
                                 charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
    return connection


def fetch_image_paths(connection, query):
    """ 从数据库中获取图片路径 """
    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()
    return results

def convert_image(input_file, output_file):
    from PIL import Image
    try:
        img = Image.open(input_file)
        img.save(output_file)
        print(f"Converted {input_file} to {output_file}")
    except IOError as e:
        print(f"Error converting {input_file}: {e}")



def write_to_excel_with_images(image_paths, output_file, base_path):
    try:
        """ 将图片路径及图片写入Excel文件 """
        wb = Workbook()
        ws = wb.active
        ws.title = "Images and Paths"

        # 写入表头
        headers = ['日期', '人员', '机组', '工作记录', '照片']
        ws.append(headers)

        # 设置列宽和行高
        ws.column_dimensions['A'].width = 20  # 设置C列的宽度
        ws.column_dimensions['D'].width = 50  # 设置C列的宽度
        ws.column_dimensions['E'].width = 20  # 设置C列的宽度
        for row in range(2, len(image_paths) + 2):
            ws.row_dimensions[row].height = 75  # 设置行高

        # 跟踪当前行号
        row_num = 2

        # 假设查询结果是一个字典列表，每个字典包含'id'和'url'键
        for item in image_paths:
            url = item['url']


            # 写入ID和路径
            ws.cell(row=row_num, column=1, value=item['update_time'])
            ws.cell(row=row_num, column=2, value=item['name'])
            ws.cell(row=row_num, column=3, value=item['title'])
            ws.cell(row=row_num, column=4, value=item['content'])
            # ws.cell(row=row_num, column=5, value=item['url'])
            # url = "data/华能融水四荣项目（中车）/014号/传动链_20230430213825.jpeg"

            if url is not None:
                path = os.path.join(base_path, url)
            else:
                row_num += 1
                continue

            if is_mpo_image(path):
                res_path = path.split(".")[0] + ".png"
                convert_image(path, res_path)
                path = res_path

            if os.path.exists(path):
                # 插入图片
                img = Image(path)
                # print(path)
                # 调整图片大小以适应单元格
                img.width = 150  # 单元格宽度
                img.height = 90  # 单元格高度

                # 设置图片的锚点
                img.anchor = f'E{row_num}'

                # 添加图片到工作表
                ws.add_image(img)
            elif path == "":
                print(f"Warning: No Image")
            else:
                print(f"Warning: Image not found at {path}")

            row_num += 1

        # from PIL import JpegImagePlugin
        # JpegImagePlugin._getmp = lambda: None
        wb.save(output_file)
    except Exception as e:
        print(f"show me: {e}")

def is_mpo_image(file_path):
    """
    检查图片文件是否为 .mpo 格式。

    :param file_path: 图片文件的路径
    :return: 如果文件是 .mpo 格式，返回 True；否则返回 False
    """
    from PIL import Image
    try:
        with Image.open(file_path) as img:
            # 获取图片格式
            format = img.format
            if format == 'MPO':
                return True
            else:
                return False
    except IOError as e:
        print(f"Error opening image: {e}")
        return False

def main(project_id, file_name):
    # 数据库连接配置
    db_config = {
        'host': '10.100.0.151',
        'port': 33060,
        'user': 'root',
        'password': 'test',
        'database': 'csp_data'
    }

    # SQL查询语句
    # sql_query = "SELECT url, path FROM project_projectunitimage LIMIT 1;"
    # sql_query = "select a.id,a.title,b.content,b.update_time,c.name,d.url from project_projectunit a LEFT JOIN project_projecttrace b on b.unit_id = a.id LEFT JOIN system_user c on c.id = b.user_id LEFT JOIN project_projectunitimage d on d.unit_id = a.id and d.trace_id = b.id where  a.project_id = " + str(
    #     project_id) + " ORDER BY a.title asc;"

    sql_query = "select a.id,a.title,b.content,b.update_time,c.name,d.url from project_projectunit a LEFT JOIN project_projecttrace b on b.unit_id = a.id LEFT JOIN system_user c on c.id = b.user_id LEFT JOIN project_projectunitimage d on d.unit_id = a.id and d.trace_id = b.id where  a.project_id = " + str(
        project_id) + " and b.update_time BETWEEN '2024-05-18 21:39:23.984440' AND '2024-08-01 21:39:23.984440' ORDER BY a.title asc limit 300,331;"



    # Excel输出文件名
    excel_output = file_name + '.xlsx'

    # 图片的基础路径
    base_image_path = 'F:' + os.sep + 'cspData'

    try:
        # 连接数据库
        conn = connect_to_database(**db_config)

        # 获取图片路径
        paths = fetch_image_paths(conn, sql_query)

        # 写入Excel
        write_to_excel_with_images(paths, excel_output, base_image_path)

        print(f"Data has been written to {excel_output}")

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if conn.open:
            conn.close()


if __name__ == '__main__':
    # listDict = [{id: 348, "name": "山东中车唐合新能源项目"}, {id: 344, "name": "三一庆阳项目"}, {id: 347, "name": "山东中车华能阿拉善项目"}]
    listDict = [{id: 373, "name": "华能融水四荣项目（中车）"}]
    for item in listDict:
        main(item.get(id), item.get("name"))
